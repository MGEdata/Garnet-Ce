from sklearn.utils import resample
import openpyxl
import numpy as np
from sklearn.model_selection import GridSearchCV
from sklearn.kernel_ridge import KernelRidge
import pandas as pd
from sklearn.preprocessing import StandardScaler
rng = np.random.RandomState(0)
np.set_printoptions(threshold=np.inf)
np.set_printoptions(suppress=True)


dataset = pd.read_csv("wavelength_training set.csv")
x = dataset.iloc[:, 0:7]
y = dataset.iloc[:, 7]

ss_x = StandardScaler()
x = ss_x.fit_transform(x)

ss_y = StandardScaler()
y = ss_y.fit_transform(y.values.reshape(-1, 1))


def model(x, y, data_predict):
    param_grid = [{"alpha": [0.14, 0.12, 0.1, 0.08, 0.06], "gamma": np.logspace(-2, 0, 10)}]
    grid = GridSearchCV(KernelRidge(kernel='rbf'), param_grid,
                        cv=10, scoring='neg_mean_absolute_error')
    grid.fit(x, y)
    rf1 = grid.best_estimator_
    model_y = rf1.fit(x, y)
    predictions = model_y.predict(data_predict)
    return (predictions)


def data_write(data, path):
    outwb = openpyxl.Workbook()
    ws = outwb.create_sheet(index=1)
    i = 1
    r = 1
    for line in data:
        for col in range(1, len(line) + 1):
            ws.cell(row=r, column=col).value = line[col - 1]
        i += 1
        r += 1
    savexlsx = path
    outwb.save(savexlsx)


dataset1 = pd.read_csv("wavelength_test set.csv")
X1 = dataset1.iloc[:, 0:7]
X22 = ss_x.transform(X1)
pre_a = np.arange(0, len(X22))
n = len(y)
m = 1000
for u in range(0, m):
    x11, y11 = resample(x, y, replace=True, n_samples=n, random_state=u)
    predict_data = model(x11, y11, X22)
    predict_y = ss_y.inverse_transform(predict_data)
    print(u)
    pre_a = np.column_stack((pre_a, predict_y))
pre_al = np.column_stack((np.mean(pre_a[:, 1:, ], axis=1), np.std(pre_a[:, 1:, ], axis=1)))


data_write(pre_al, "111.xlsx")
